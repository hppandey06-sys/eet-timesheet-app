"""
Excel export — generates per-member, per-project, per-month timesheet
in the DC Power format.
"""

from io import BytesIO
from datetime import date, timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar


def export_member_timesheet(member, project_name, project_acts, entries,
                            from_date, to_date, company=None):
    """
    Generate Excel timesheet matching the DC Power format.

    Args:
        member: dict with 'name', 'company', 'discipline'
        project_name: e.g., 'DC Power' or 'ALL PROJECTS' for combined view
        project_acts: list of dicts with 'code', 'description'
        entries: list of entries for this member + date range
        from_date, to_date: date range
        company: org name (optional, defaults to member's company)
    """
    is_all_projects = (project_name == 'ALL PROJECTS')

    wb = Workbook()
    wb.remove(wb.active)

    # Group dates by month
    months = {}
    cur = date.fromisoformat(from_date) if isinstance(from_date, str) else from_date
    end = date.fromisoformat(to_date) if isinstance(to_date, str) else to_date

    while cur <= end:
        ym = cur.strftime("%Y-%m")
        if ym not in months:
            months[ym] = []
        months[ym].append(cur)
        cur += timedelta(days=1)

    org_name = company or member.get("company", "GCC")
    member_name = member.get("name", "Unknown")

    # Borders and styles
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1F4E79")
    weekend_fill = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="FFE699")
    total_font = Font(bold=True, size=11)
    project_section_fill = PatternFill("solid", fgColor="D9E1F2")

    for ym, dates in months.items():
        month_label = dates[0].strftime("%b %y")
        ws = wb.create_sheet(month_label)

        # Filter entries to this month
        date_strs = [d.strftime("%Y-%m-%d") for d in dates]
        month_entries = [e for e in entries if e.get("entry_date") in date_strs]

        # Title
        title = f"Timesheet for {project_name}" if not is_all_projects else "Timesheet — All Projects"
        ws["A1"] = title
        ws["A1"].font = title_font
        ws.merge_cells(f"A1:{get_column_letter(len(dates) + 1)}1")

        ws["A2"] = f"Timesheet - {member_name}."
        ws["A2"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A2:{get_column_letter(len(dates) + 1)}2")

        ws["A3"] = "Month:"
        ws["A3"].font = Font(bold=True)
        ws["B3"] = dates[0].replace(day=1)
        ws["B3"].number_format = "mmm-yy"

        ws["A4"] = "Org. Name"
        ws["A4"].font = Font(bold=True)
        ws["B4"] = org_name

        # Headers row
        ws["A6"] = "Hours Worked (max. 8hrs per day)"
        ws["A6"].font = Font(italic=True, size=10, color="606060")
        ws.merge_cells(f"A6:{get_column_letter(len(dates) + 1)}6")

        # Date row
        ws["A7"] = "Date:"
        ws["A7"].font = header_font
        ws["A7"].fill = header_fill
        ws["A7"].border = border
        ws["A7"].alignment = Alignment(horizontal="center", vertical="center")

        for i, d in enumerate(dates):
            cell = ws.cell(row=7, column=i + 2)
            cell.value = d
            cell.number_format = "dd-mmm"
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i + 2)].width = 9

        # Day name row
        ws["A8"] = ""
        ws["A8"].fill = header_fill
        ws["A8"].border = border
        for i, d in enumerate(dates):
            cell = ws.cell(row=8, column=i + 2)
            cell.value = d.strftime("%a")
            cell.font = Font(italic=True, size=9, color="606060")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if d.weekday() >= 5:
                cell.fill = weekend_fill

        ws.column_dimensions["A"].width = 32

        if is_all_projects:
            row_num = render_all_projects(
                ws, month_entries, dates, project_section_fill, border, weekend_fill, total_fill, total_font
            )
        else:
            # Single project - filter to that project
            proj_entries = [e for e in month_entries if e.get("proj") == project_name]
            row_num = render_single_project(
                ws, proj_entries, project_acts, dates, project_section_fill,
                border, weekend_fill, total_fill, total_font, project_name, start_row=9
            )

        # Final grand total row
        total_row = row_num
        ws.cell(row=total_row, column=1).value = "GRAND TOTAL"
        ws.cell(row=total_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(row=total_row, column=1).border = border
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")

        for i, d in enumerate(dates):
            ds = d.strftime("%Y-%m-%d")
            day_total = sum(float(e.get("hrs", 0)) for e in month_entries if e.get("entry_date") == ds)
            cell = ws.cell(row=total_row, column=i + 2)
            cell.value = day_total if day_total > 0 else None
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Grand total of all hours
        grand_total = sum(float(e.get("hrs", 0)) for e in month_entries)
        gt_col = len(dates) + 2
        cell = ws.cell(row=total_row, column=gt_col)
        cell.value = grand_total if grand_total > 0 else None
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(gt_col)].width = 10

        # Total label header
        ws.cell(row=7, column=gt_col).value = "Total"
        ws.cell(row=7, column=gt_col).font = header_font
        ws.cell(row=7, column=gt_col).fill = header_fill
        ws.cell(row=7, column=gt_col).border = border
        ws.cell(row=7, column=gt_col).alignment = Alignment(horizontal="center")

        # Freeze panes
        ws.freeze_panes = "B9"

    # Output buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def render_single_project(ws, proj_entries, project_acts, dates,
                          section_fill, border, weekend_fill, total_fill, total_font,
                          project_name=None, start_row=9):
    """Render rows for a single project. Returns next available row."""
    if project_name:
        cell = ws.cell(row=start_row, column=1)
        cell.value = project_name
        cell.font = Font(bold=True, color="1F4E79", size=11)
        cell.fill = section_fill
        cell.border = border
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="left")
        # Fill section row
        for i in range(len(dates) + 1):
            c = ws.cell(row=start_row, column=i + 2)
            c.fill = section_fill
            c.border = border
        start_row += 1

    if not project_acts:
        # Use activity codes from actual entries
        used_acts = sorted(set(e.get("act") for e in proj_entries))
        project_acts = [{"code": code, "description": ""} for code in used_acts] or [{"code": "OTHERS", "description": "Other"}]

    row_num = start_row
    for act_info in project_acts:
        code = act_info["code"]
        desc = act_info.get("description", "")
        label = f"{code} - {desc}" if desc else code

        cell = ws.cell(row=row_num, column=1)
        cell.value = label
        cell.font = Font(size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Fill hours per day
        row_total = 0
        for i, d in enumerate(dates):
            ds = d.strftime("%Y-%m-%d")
            hrs_for_day = sum(
                float(e.get("hrs", 0)) for e in proj_entries
                if e.get("entry_date") == ds and e.get("act") == code
            )
            cell = ws.cell(row=row_num, column=i + 2)
            cell.value = hrs_for_day if hrs_for_day > 0 else None
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if d.weekday() >= 5:
                cell.fill = weekend_fill
            row_total += hrs_for_day

        # Row total at end
        cell = ws.cell(row=row_num, column=len(dates) + 2)
        cell.value = row_total if row_total > 0 else None
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

        row_num += 1

    # Subtotal row for this project
    subtotal_row = row_num
    ws.cell(row=subtotal_row, column=1).value = "Subtotal"
    ws.cell(row=subtotal_row, column=1).font = total_font
    ws.cell(row=subtotal_row, column=1).fill = total_fill
    ws.cell(row=subtotal_row, column=1).border = border

    for i, d in enumerate(dates):
        col_letter = get_column_letter(i + 2)
        cell = ws.cell(row=subtotal_row, column=i + 2)
        cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{subtotal_row - 1})"
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Subtotal end column
    cell = ws.cell(row=subtotal_row, column=len(dates) + 2)
    cell.value = f"=SUM(B{subtotal_row}:{get_column_letter(len(dates) + 1)}{subtotal_row})"
    cell.font = total_font
    cell.fill = total_fill
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

    return subtotal_row + 2  # leave gap before next project


def render_all_projects(ws, entries, dates, section_fill, border, weekend_fill, total_fill, total_font):
    """Render rows grouped by project."""
    # Get unique projects from entries
    projects_used = sorted(set(e.get("proj") for e in entries if e.get("proj")))

    if not projects_used:
        ws.cell(row=9, column=1).value = "(No entries in this month)"
        ws.cell(row=9, column=1).font = Font(italic=True, color="808080")
        return 11

    row_num = 9
    for proj_name in projects_used:
        # Get unique activities used for this project
        proj_entries = [e for e in entries if e.get("proj") == proj_name]
        used_acts = sorted(set(e.get("act") for e in proj_entries))
        project_acts = [{"code": code, "description": ""} for code in used_acts]

        # Render this project
        row_num = render_single_project(
            ws, proj_entries, project_acts, dates, section_fill,
            border, weekend_fill, total_fill, total_font,
            project_name=proj_name, start_row=row_num
        )

    return row_num
